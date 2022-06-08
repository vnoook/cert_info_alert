# Программа для проверки дат окончания сертификатов.
# В папке с программой должны быть папки {dir_cers} и {dir_txts}.
# В папку {dir_cers} скопируйте сертификаты с расширение ".cer".
# Папка {dir_txts} нужна для хранения дампов сертификатов из {dir_cers}.
# При наличии сертификатов в папке {dir_cers} с помощью программы Windows "certutil"
# создаются дампы с текстовом формате.
# Эти дампы анализируются, и сортируются по порядку изложенному в переменной {tuple_search_string}
# и выгружаются в xlsx файл. В полученном файле xlsx подсвечиваются ячейки дат со скором окончанием.
# Красным подсвечиваются сроки "месяц до окончания" - 30 дней, "розовым" полтора месяца - 45 дней,
# "зелёным" более 45 дней, "серым" просроченные сертификаты от текущей даты.
# ...
# INSTALL
# pip install openpyxl
# COMPILE
# pyinstaller -F cert_info_alert.py
# ...

# TODO
# 199, 224

import os
import datetime
import subprocess
import openpyxl
import openpyxl.utils
import openpyxl.styles

# переменные
etx_txt = '.txt'  # расширение файлов с текстом
etx_cer = '.cer'  # расширение файлов сертификатов
dir_cers = r'cer_s'  # папка для сохранения сертификатов
dir_txts = r'txt_s'  # папка для дампов сертификатов
# команда для командной строки 'for /r {dir_cers} %i in (*.cer) do certutil "%i" > "{dir_txts}\%~ni.txt"'
cer_command = 'certutil.exe "path_cer" > "path_txt"'
# файл шаблон для называния файла выгрузки
name_file_xlsx = 'cert.xlsx'
# конечный список со строками данных в нужном порядке который выгружается в эксель
list_of_strings_from_files = []
# строка для заполнения ячеек с данными из "нестандартных" сертификатов
value_empty_string = 'нет данных в дампе сертификата'


# проверка на существование папок ({dir_cers}, {dir_txts}) для сертификатов и дампов
# и создание их, если их нет
def check_exists_dirs():
    # переход в корневую папку
    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    # проверка на существование папок
    if not os.path.exists(dir_cers):
        os.mkdir(dir_cers)
    if not os.path.exists(dir_txts):
        os.mkdir(dir_txts)


# функция очищающая папку {dir_txts} для создания новых дампов сертификатов
def clean_dir_txts():
    # переход в папку дампов
    os.chdir(os.path.join(os.path.dirname(os.path.realpath(__file__)), dir_txts))

    # поиск файлов txt и их удаление
    for data_of_scan in os.scandir():
        if data_of_scan.is_file() and os.path.splitext(os.path.split(data_of_scan)[1])[1] == etx_txt:
            os.remove(data_of_scan)
    print('\n(1)...папка от старых дампов сертификатов очищена')


# функция создающая дампы файлов и складывающая их в папку {dir_txts}
def do_txt_from_cer():
    # переход в корневую папку
    os.chdir(os.path.dirname(os.path.realpath(__file__)))
    # переход в папку с сертификатами
    # os.chdir(os.path.join(os.path.dirname(os.path.realpath(__file__)), dir_cers))

    for data_of_scan in os.scandir(os.path.join(os.path.dirname(os.path.realpath(__file__)), dir_cers)):
        if data_of_scan.is_file() and os.path.splitext(os.path.split(data_of_scan)[1])[1] == etx_cer:
            path_cer = os.path.join(dir_cers, data_of_scan.name)
            path_txt = os.path.join(dir_txts, str(data_of_scan.name).replace('.cer','.txt'))
            subprocess.run(cer_command.replace('path_cer', path_cer).replace('path_txt', path_txt),
                           stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL, shell=True, encoding='utf-8')

    print('\n(2)...дампы сертификатов вновь созданы')


# функция чтения, фильтрации и сортировки дампов сертификатов
# а также формирования конечной таблицы для вывода её в xlsx
def processing_txt_files():
    # переход в папку с дампами
    os.chdir(os.path.join(os.path.dirname(os.path.realpath(__file__)), dir_txts))

    # поля и их порядок для поиска в дампах
    # если добавить или удалить поля (кроме последнего), то алгоритм не собъётся
    tuple_search_string = (
                           'SN',  # фамилия 'SN='
                           'G',   # имя отчество 'G='
                           'NotAfter',   # дата конца 'NotAfter:'
                           'NotBefore',  # дата начала 'NotBefore:'
                           'CN',  # организация выдавшая сертификат 'CN='
                           'Хеш сертификата(sha1)',  # отпечаток 'Хеш сертификата(sha1):'
                           'Серийный номер',  # отличие между органами выдавшими сертификат
                           'полный путь до дампа'  # полный путь до дампа
                          )
    # чтение и обработка каждого файла в список, и добавление в конец списка "путь к дампу"
    for data_of_scan in os.scandir():
        if data_of_scan.is_file() and os.path.splitext(os.path.split(data_of_scan)[1])[1] == etx_txt:
            # получил все строки из файла
            all_strings_from_file = []
            with open(data_of_scan.name, 'r') as txt_file:
                all_strings_from_file = txt_file.read().splitlines()

            # выбрал из всех строк те, которые имеются суфиксы из tuple_search_string
            # и заменил в строках где есть первый слева символ "=" на ":", для простоты в будущем
            list_of_need_strings = []
            for string_from_file in all_strings_from_file:
                string_from_file = string_from_file.strip()
                if (string_from_file.split(':', maxsplit=1)[0] in tuple_search_string) or\
                   (string_from_file.split('=', maxsplit=1)[0] in tuple_search_string):
                    list_of_need_strings.append(string_from_file.replace('=', ':', 1))

            # из списка list_of_need_strings нужно вычислить задвоенные суфиксы
            # И создать порядок для формирования конечного списка для выгрузки в xlsx
            list_of_need_strings_sorted = []
            for suffix in tuple_search_string:

                # счётчик количества повторений суфиксов
                # если не ноль, то в списке есть повторения, которые надо игнорировать
                count_suffix = 0
                for string_from_need_list in list_of_need_strings:
                    suffix_of_need_string = string_from_need_list.split(':', maxsplit=1)[0]

                    if suffix_of_need_string == suffix:
                        value_of_string = string_from_need_list.split(':', maxsplit=1)[1]
                        # если счётчик равен 0, то это первое повторение суфикса, остальные будут игнорироваться

                        if count_suffix == 0:
                            # индивидуальные обработки колонок

                            if suffix_of_need_string == 'Хеш сертификата(sha1)':
                                # удаление пробелов в Хеше
                                list_of_need_strings_sorted.append(value_of_string.replace(' ', ''))
                            elif suffix_of_need_string in ('NotAfter', 'NotBefore'):
                                date_val = value_of_string.strip().split(' ', maxsplit=1)[0].replace('.', '-')
                                list_of_need_strings_sorted.append(datetime.datetime.strptime(date_val, '%d-%m-%Y'))
                            else:
                                list_of_need_strings_sorted.append(value_of_string.strip())
                        count_suffix += 1

                # если за все проходы строк не найден суфикс, то вставить "заглушку"
                if count_suffix == 0:
                    list_of_need_strings_sorted.append(value_empty_string)

            # добавил в последний индекс ссылку на дамп сертификата
            list_of_need_strings_sorted[-1] = os.path.abspath(data_of_scan)

            # создал список списков
            list_of_strings_from_files.append(list_of_need_strings_sorted)

    # добавил в первую строку шапку из названий колонок
    list_of_strings_from_files.insert(0, list(val for val in tuple_search_string))

    print('\n(3)...дампы сертификатов прочитаны и таблица для записи в xlsx готова')


# перенос списка данных в файл xlsx
def do_xlsx():
    # переход в корневую папку
    os.chdir(os.path.dirname(os.path.realpath(__file__)))

    # вычисление текущей даты в формате дд-мм-гггг
    today_date = datetime.datetime.now().date()

    # стиль для ячеек с -1 днями
    style_1 = openpyxl.styles.NamedStyle(name='style_1')
    style_1.fill = openpyxl.styles.PatternFill('solid', fgColor='00C0C0C0')   # серый
    # стиль для ячеек между 0-30 днями
    style_30 = openpyxl.styles.NamedStyle(name='style_30')
    style_30.fill = openpyxl.styles.PatternFill('solid', fgColor='00FF0000')  # красный
    # стиль для ячеек между 30-45 днями
    style_45 = openpyxl.styles.NamedStyle(name='style_45')
    style_45.fill = openpyxl.styles.PatternFill('solid', fgColor='00FF99CC')  # розовый
    # стиль для ячеек с более 45 днями
    style_46 = openpyxl.styles.NamedStyle(name='style_46')
    style_46.fill = openpyxl.styles.PatternFill('solid', fgColor='0099CC00')  # зелёный

    # создание xlsx файла
    file_xlsx = openpyxl.Workbook()
    file_xlsx_s = file_xlsx.active

    # вычисляются "высота" и "длина" данных
    row_of_list = len(list_of_strings_from_files)
    col_of_list = len(list_of_strings_from_files[0])

    # заполнение ячеек значениями с предварительной их обработкой
    # а также подсветка в зависимости от значения
    if row_of_list > 0:
        for col in range(1, col_of_list+1):
            max_len_value_of_col = 0
            for row in range(1, row_of_list+1):
                value_of_string_for_cell = list_of_strings_from_files[row-1][col-1]

                # если ячейки с датами, то подсветить
                if file_xlsx_s.cell(1, col).value == 'NotAfter':
                    if value_of_string_for_cell != value_empty_string:
                        print()
                        print(f'{value_of_string_for_cell = } ... {value_empty_string = }')
                        # print(f'{datetime.datetime.date(value_of_string_for_cell) = } ... {list_of_strings_from_files[row-1] = }')
                        # print(f'{datetime.datetime.date(value_of_string_for_cell) = } ... {list_of_strings_from_files[row] = }')
                        print()

                        # дата из сертификата
                        cert_date = datetime.datetime.date(value_of_string_for_cell)
                        # разница между датой из сертификата и текущей
                        delta_date = cert_date - today_date

                        # распределение по стилям разных значений разниц дат
                        if delta_date <= datetime.timedelta(0):
                            file_xlsx_s.cell(row, col).style = style_1
                        elif (delta_date > datetime.timedelta(0)) and (delta_date <= datetime.timedelta(30)):
                            file_xlsx_s.cell(row, col).style = style_30
                        elif (delta_date > datetime.timedelta(30)) and (delta_date <= datetime.timedelta(45)):
                            file_xlsx_s.cell(row, col).style = style_45
                        elif delta_date > datetime.timedelta(45):
                            file_xlsx_s.cell(row, col).style = style_46

                        file_xlsx_s.cell(row, col, datetime.datetime.date(value_of_string_for_cell))

                elif file_xlsx_s.cell(1, col).value == 'NotBefore':
                    if value_of_string_for_cell != value_empty_string:
                        file_xlsx_s.cell(row, col, datetime.datetime.date(value_of_string_for_cell))

                # в строке пути к дампу добавляется ссылка путём на сертификат
                elif file_xlsx_s.cell(1, col).value == 'полный путь до дампа':
                    file_xlsx_s.cell(row, col).hyperlink =\
                        value_of_string_for_cell.replace('.txt', '.cer').replace(dir_txts, dir_cers).strip()
                    file_xlsx_s.cell(row, col).value = value_of_string_for_cell.strip()

                else:
                    file_xlsx_s.cell(row, col, value_of_string_for_cell.strip())

                # вычисление самого длинного значения в колонке
                if len(str(file_xlsx_s.cell(row, col).value)) > max_len_value_of_col:
                    max_len_value_of_col = len(str(file_xlsx_s.cell(row, col).value).strip())

            # установка ширины ячеек по всем колонкам для красоты в экселе
            file_xlsx_s.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len_value_of_col * 1.1
            max_len_value_of_col = 0

    # включение фильтра
    file_xlsx_s.auto_filter.ref = 'A1:' + openpyxl.utils.get_column_letter(col_of_list)+'1'

    # сохраняю файл xlsx с добавлением в имя текущей даты
    file_xlsx.save(name_file_xlsx.replace('cert', 'cert_'+str(datetime.datetime.date(datetime.datetime.now()))))
    # закрываю файл
    file_xlsx.close()

    print('\n(4)...файл с данными сертификатов собран')
    print('\n(5)...ГОТОВО!')
    input('\nнажмите ENTER')


def run():
    check_exists_dirs()
    clean_dir_txts()
    do_txt_from_cer()
    processing_txt_files()
    do_xlsx()


if __name__ == '__main__':
    run()
